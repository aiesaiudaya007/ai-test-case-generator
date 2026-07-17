"""Export scenarios + test cases to Excel (for TestRail/Zephyr/Xray import) and JSON."""
import json
import logging
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import Scenario, Story, TestCase

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def export_to_excel(story: Story, scenarios: List[Scenario], test_cases: List[TestCase], output_path: str) -> None:
    wb = Workbook()

    _write_scenarios_sheet(wb.active, story, scenarios)
    wb.active.title = "Scenarios"
    _write_test_cases_sheet(wb.create_sheet("Test Cases"), story, test_cases)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info("Excel written to %s", output_path)


def _write_scenarios_sheet(ws, story: Story, scenarios: List[Scenario]) -> None:
    ws.append([f"Scenarios for {story.key}: {story.summary}"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])

    headers = ["Scenario ID", "Title", "Category", "Description", "Related AC"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for sc in scenarios:
        ws.append([sc.id, sc.title, sc.category, sc.description, sc.related_ac])

    widths = [12, 32, 14, 50, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=4):
        for cell in row:
            cell.alignment = WRAP


def _write_test_cases_sheet(ws, story: Story, test_cases: List[TestCase]) -> None:
    headers = [
        "Test Case ID", "Scenario ID", "Scenario", "Title", "Type", "Priority",
        "Preconditions", "Steps", "Test Data", "Expected Result",
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for tc in test_cases:
        steps_text = "\n".join(f"{i}. {s}" for i, s in enumerate(tc.steps, start=1))
        ws.append([
            tc.test_case_id, tc.scenario_id, tc.scenario_title, tc.title, tc.type,
            tc.priority, tc.preconditions, steps_text, tc.test_data, tc.expected_result,
        ])

    widths = [14, 12, 30, 30, 10, 10, 28, 45, 28, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP


def export_to_json(story: Story, scenarios: List[Scenario], test_cases: List[TestCase], output_path: str) -> None:
    data = {
        "story": story.as_dict(),
        "scenarios": [s.as_dict() for s in scenarios],
        "test_cases": [t.as_dict() for t in test_cases],
    }
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w") as f:
        json.dump(data, f, indent=2)
    logger.info("JSON written to %s", output_path)
